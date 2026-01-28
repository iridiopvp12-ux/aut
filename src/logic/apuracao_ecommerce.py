import logging
import shutil
import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from typing import List

# ==============================================================================
# 1. FUNÇÕES AUXILIARES E ESTILO
# ==============================================================================

def _ler_valor_mesclado(ws: Worksheet, linha: int, coluna: int):
    """Lê valor lidando com merge."""
    cell = ws.cell(row=linha, column=coluna)
    if cell.value is None:
        for merged_range in ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                top_left_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                return top_left_cell.value
    return cell.value

def _limpar_cfop_excel(valor_celula) -> List[str]:
    if not valor_celula: return []
    s = str(valor_celula)
    if isinstance(valor_celula, float) and s.endswith('.0'): s = s[:-2]
    s = s.replace(' ', '').strip()
    partes = s.split('/')
    return [p for p in partes if p.isdigit()]

def _escrever_seguro(ws: Worksheet, linha: int, coluna: int, valor: float):
    cell = ws.cell(row=linha, column=coluna)
    if isinstance(cell, MergedCell):
        for merged_range in ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                top_left_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                top_left_cell.value = valor
                return cell
    else:
        cell.value = valor
        return cell

def _aplicar_estilo_tabela_sobras(ws, linha, col_inicial, col_final, is_header=False, cor_header="4F81BD"):
    """Estilização do relatório lateral."""
    thick = Side(border_style="medium", color="000000")
    thin = Side(border_style="thin", color="000000")
    borda = Border(top=thin, left=thin, right=thin, bottom=thin)

    for c in range(col_inicial, col_final + 1):
        cell = ws.cell(row=linha, column=c)
        cell.border = borda
        if is_header:
            cell.fill = PatternFill(start_color=cor_header, end_color=cor_header, fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            if c == col_final: cell.alignment = Alignment(horizontal="left")
            elif c in [col_inicial, col_inicial+1]: cell.alignment = Alignment(horizontal="center")
            else:
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = '#,##0.00'

def _gerar_relatorio_sobras(ws: Worksheet, df: pd.DataFrame, col_inicio: int, titulo_bloco: str, cor_fundo: str):
    """Gera o relatório de itens não utilizados na lateral."""
    if df is None or df.empty: return

    logging.info(f"[E-COMMERCE] Gerando relatório de sobras: {titulo_bloco}")

    mask_sobra = (~df['Utilizado']) & (df['Total Operação'] > 0.01)
    df_sobra = df.loc[mask_sobra].copy().sort_values(by='Total Operação', ascending=False)

    if df_sobra.empty: return

    C_CFOP = col_inicio
    C_ALIQ = col_inicio + 1
    C_VALOR = col_inicio + 2
    C_BASE = col_inicio + 3
    C_ICMS = col_inicio + 4
    C_MOTIVO = col_inicio + 5

    LINHA = 5

    # Título
    ws.merge_cells(start_row=LINHA-2, start_column=col_inicio, end_row=LINHA-2, end_column=C_MOTIVO)
    cell_title = ws.cell(row=LINHA-2, column=col_inicio, value=f"⚠️ SOBRAS - {titulo_bloco}")
    cell_title.font = Font(bold=True, color="FFFFFF", size=11)
    cell_title.fill = PatternFill(start_color=cor_fundo, end_color=cor_fundo, fill_type="solid")
    cell_title.alignment = Alignment(horizontal="center")

    # Cabeçalho da Tabela
    titulos = {C_CFOP: "CFOP", C_ALIQ: "Aliq %", C_VALOR: "Vlr Contábil", C_BASE: "Base Calc", C_ICMS: "ICMS", C_MOTIVO: "Provável Motivo"}
    for col, titulo in titulos.items(): _escrever_seguro(ws, LINHA-1, col, titulo)
    _aplicar_estilo_tabela_sobras(ws, LINHA-1, col_inicio, C_MOTIVO, is_header=True, cor_header=cor_fundo)

    # Dados
    for _, row in df_sobra.iterrows():
        if LINHA > 200: break

        cfop = str(row['CFOP (SPED)']).replace('.0', '')
        aliq = row['Alíquota (SPED)']

        motivo = "Não mapeado"
        if aliq == 0: motivo = "Alíquota Zero"
        elif cfop in ['1403', '2403', '5403', '6403']: motivo = "ST (Aliq 0)"
        elif cfop.startswith('59') or cfop.startswith('69'): motivo = "Remessa/Isento"

        _escrever_seguro(ws, LINHA, C_CFOP, cfop)
        _escrever_seguro(ws, LINHA, C_ALIQ, aliq)
        _escrever_seguro(ws, LINHA, C_VALOR, row['Total Operação'])
        _escrever_seguro(ws, LINHA, C_BASE, row['Base de Cálculo ICMS'])
        _escrever_seguro(ws, LINHA, C_ICMS, row['Total ICMS'])
        _escrever_seguro(ws, LINHA, C_MOTIVO, motivo)

        _aplicar_estilo_tabela_sobras(ws, LINHA, col_inicio, C_MOTIVO, is_header=False)
        LINHA += 1

def _preparar_dataframe(df_orig: pd.DataFrame) -> pd.DataFrame:
    if df_orig is None or df_orig.empty:
        return pd.DataFrame()

    df = df_orig.copy()
    df['Utilizado'] = False

    if 'CFOP (SPED)' in df.columns:
        df['CFOP (SPED)'] = df['CFOP (SPED)'].apply(
            lambda x: str(int(x)) if pd.notnull(x) and isinstance(x, (int, float)) else str(x).strip()
        )

    cols_num = ['Total Operação', 'Base de Cálculo ICMS', 'Total ICMS', 'Alíquota (SPED)']
    for col in cols_num:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0.0)
        else:
            df[col] = 0.0

    return df

# ==============================================================================
# 2. LÓGICA MISTA INTELIGENTE (ENTRADAS E SAÍDAS NO MESMO QUADRO)
# ==============================================================================

def preencher_quadro_misto_ecommerce(ws: Worksheet, df_entradas: pd.DataFrame, df_saidas: pd.DataFrame):
    logging.info("[E-COMMERCE] Iniciando preenchimento HÍBRIDO (Entradas + Saídas)...")

    df_ent = _preparar_dataframe(df_entradas)
    df_sai = _preparar_dataframe(df_saidas)

    def processar_range_inteligente(inicio, fim):
        for linha in range(inicio, fim + 1):
            valor_cfop_excel = _ler_valor_mesclado(ws, linha, 2)
            lista_cfops = _limpar_cfop_excel(valor_cfop_excel)

            if not lista_cfops: continue

            primeiro_digito = lista_cfops[0][0]

            df_alvo = None
            if primeiro_digito in ['1', '2', '3']:
                df_alvo = df_ent
            elif primeiro_digito in ['5', '6', '7']:
                df_alvo = df_sai

            if df_alvo is None or df_alvo.empty:
                continue

            mask = df_alvo['CFOP (SPED)'].isin(lista_cfops)

            if mask.any():
                df_alvo.loc[mask, 'Utilizado'] = True

            df_filtered = df_alvo.loc[mask]

            if not df_filtered.empty:
                soma_base = df_filtered['Base de Cálculo ICMS'].sum()
                soma_icms = df_filtered['Total ICMS'].sum()

                if soma_base > 0: _escrever_seguro(ws, linha, 3, soma_base)
                if soma_icms > 0: _escrever_seguro(ws, linha, 4, soma_icms)

    # --- Executa as faixas solicitadas ---
    processar_range_inteligente(9, 15)
    processar_range_inteligente(20, 27)
    processar_range_inteligente(32, 34)
    processar_range_inteligente(39, 40)

    # --- Preenchimento dos Totalizadores (Células Fixas) ---
    if not df_ent.empty:
        total_icms_ent = df_ent['Total ICMS'].sum()
        if total_icms_ent > 0:
            _escrever_seguro(ws, 62, 5, total_icms_ent) # E62
            _escrever_seguro(ws, 50, 3, total_icms_ent) # C50

    if not df_sai.empty:
        total_icms_sai = df_sai['Total ICMS'].sum()
        if total_icms_sai > 0:
            _escrever_seguro(ws, 54, 5, total_icms_sai) # E54

    # --- RELATÓRIOS LATERAIS ---
    # ATENÇÃO: Removido o relatório de ENTRADAS conforme solicitado.

    # Saídas na Coluna 16 (P) - Agora tem espaço pois tiramos as entradas
    if not df_sai.empty:
        _gerar_relatorio_sobras(ws, df_sai, 16, "SAÍDAS", "C65911")   # Laranja

# ==============================================================================
# 3. FUNÇÃO PRINCIPAL (ORQUESTRADOR)
# ==============================================================================

def preencher_template_ecommerce(template_path: Path, df_entradas: pd.DataFrame, df_saidas: pd.DataFrame = None) -> str:
    logging.info(f"[E-COMMERCE] Processando arquivo base: {template_path}")

    if (df_entradas is None or df_entradas.empty) and (df_saidas is None or df_saidas.empty):
        return str(template_path)

    try:
        path_origem = Path(template_path)
        pasta = path_origem.parent
        novo_nome = f"{path_origem.stem}_ECOMMERCE_PREENCHIDA{path_origem.suffix}"
        path_destino = pasta / novo_nome

        shutil.copy(path_origem, path_destino)
        wb = load_workbook(path_destino)
        ws = wb["Entradas"] if "Entradas" in wb.sheetnames else wb.active

        preencher_quadro_misto_ecommerce(ws, df_entradas, df_saidas)

        wb.save(path_destino)
        logging.info(f"[E-COMMERCE] Sucesso! Arquivo gerado: {path_destino}")

        return str(path_destino)

    except Exception as e:
        logging.error(f"[E-COMMERCE] Erro fatal: {e}")
        raise e