import logging
import pandas as pd
from pathlib import Path
from typing import Dict

# --- IMPORTAÇÕES DO OPENPYXL ---
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

def gerar_relatorio_excel(
    caminho_saida: Path,
    df_recon_relatorio: pd.DataFrame,
    df_itens_aba: pd.DataFrame,
    df_aliquota_aba: pd.DataFrame,
    df_totalizadores_entrada: pd.DataFrame,
    df_totalizadores_saida: pd.DataFrame,
    df_cte_bruto_aba: pd.DataFrame
) -> None:
    """Gera o arquivo Excel final com todas as abas e formatações."""

    writer = None
    try:
        writer = pd.ExcelWriter(str(caminho_saida), engine='openpyxl')
    except ImportError:
        msg = "'openpyxl' é necessário. Instale com: pip install openpyxl"
        logging.error(msg)
        raise ImportError(msg)

    # --- Definição dos Estilos (Sintaxe OpenPyXL) ---
    header_fill = PatternFill(start_color='2D3E50', end_color='2D3E50', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ok_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    ok_font = Font(color='006100')
    divergent_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    divergent_font = Font(color='9C0006')
    revisar_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    revisar_font = Font(color='9C6500')
    multiple_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    multiple_font = Font(bold=True)
    na_font = Font(color='808080', italic=True)

    format_currency = 'R$ #,##0.00'
    format_percent = '0.00%'
    format_number = '#,##0.0000'
    format_mva = '0.00'
    format_aliquota = '0.00'

    def apply_styles_and_rules_v2(ws: Worksheet, df: pd.DataFrame, status_cols_map: Dict, cfop_cols_map: Dict, col_formats_map: Dict):
        """Aplica cabeçalho, larguras, formatos e formatação condicional."""

        if df.empty: return

        max_row = ws.max_row
        max_col = ws.max_column

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            cell.border = thin_border

        for col_name, col_idx in status_cols_map.items():
            col_letter = get_column_letter(col_idx + 1)
            cell_range = f"{col_letter}2:{col_letter}{max_row}"
            first_cell = f"{col_letter}2"

            if col_name == 'TIPO_NOTA':
                ws.conditional_formatting.add(cell_range, CellIsRule(operator='equal', formula=['"Devolução"'], stopIfTrue=True, fill=divergent_fill, font=divergent_font))
                ws.conditional_formatting.add(cell_range, CellIsRule(operator='equal', formula=['"Complementar"'], stopIfTrue=True, fill=revisar_fill, font=revisar_font))
                ws.conditional_formatting.add(cell_range, CellIsRule(operator='equal', formula=['"Ajuste"'], stopIfTrue=True, fill=revisar_fill, font=revisar_font))
                ws.conditional_formatting.add(cell_range, FormulaRule(formula=[f'ISNUMBER(SEARCH("Energia Elétrica",{first_cell}))'], stopIfTrue=True, fill=revisar_fill, font=revisar_font))
                ws.conditional_formatting.add(cell_range, FormulaRule(formula=[f'ISNUMBER(SEARCH("Comunicação",{first_cell}))'], stopIfTrue=True, fill=revisar_fill, font=revisar_font))

            ws.conditional_formatting.add(cell_range, CellIsRule(operator='equal', formula=['"DIVERGENTE"'], stopIfTrue=True, fill=divergent_fill, font=divergent_font))
            ws.conditional_formatting.add(cell_range, CellIsRule(operator='equal', formula=['"FALTA XML"'], stopIfTrue=True, fill=divergent_fill, font=divergent_font))
            ws.conditional_formatting.add(cell_range, CellIsRule(operator='equal', formula=['"FALTA NO SPED"'], stopIfTrue=True, fill=divergent_fill, font=divergent_font))

            ws.conditional_formatting.add(cell_range, FormulaRule(formula=[f'ISNUMBER(SEARCH("REVISAR",{first_cell}))'], stopIfTrue=True, fill=revisar_fill, font=revisar_font))
            ws.conditional_formatting.add(cell_range, CellIsRule(operator='equal', formula=['"SEM CNPJ NO XML"'], stopIfTrue=True, fill=revisar_fill, font=revisar_font))

            ws.conditional_formatting.add(cell_range, CellIsRule(operator='equal', formula=['"OK"'], stopIfTrue=True, fill=ok_fill, font=ok_font))

            ws.conditional_formatting.add(cell_range, CellIsRule(operator='equal', formula=['"N/A"'], stopIfTrue=True, font=na_font))

        for col_name, col_idx in cfop_cols_map.items():
            col_letter = get_column_letter(col_idx + 1)
            cell_range = f"{col_letter}2:{col_letter}{max_row}"
            first_cell = f"{col_letter}2"
            ws.conditional_formatting.add(cell_range, FormulaRule(formula=[f'ISNUMBER(SEARCH("/",{first_cell}))'], stopIfTrue=True, fill=multiple_fill, font=multiple_font))
            if col_name == 'STATUS_CFOP':
                ws.conditional_formatting.add(cell_range, FormulaRule(formula=[f'ISNUMBER(SEARCH("Múltiplos",{first_cell}))'], stopIfTrue=True, fill=multiple_fill, font=multiple_font))

        for col_idx_0based, (col_name, width, num_format) in col_formats_map.items():
            col_letter = get_column_letter(col_idx_0based + 1)
            ws.column_dimensions[col_letter].width = width

            if num_format and max_row > 1:
                for row_idx in range(2, max_row + 1):
                    ws.cell(row=row_idx, column=col_idx_0based + 1).number_format = num_format

        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions

    try:
        # --- GERAÇÃO DA ABA 'Conciliacao' ---
        if not df_recon_relatorio.empty:
            df_recon_relatorio.to_excel(writer, sheet_name='Conciliacao', index=False)
            ws_conciliacao = writer.sheets['Conciliacao']

            status_cols_conc = {}
            cfop_cols_conc = {}
            col_formats_conc = {}

            for col_idx, col_name in enumerate(df_recon_relatorio.columns):
                num_format_to_apply = None
                width = 18

                if col_name.startswith('STATUS_') or col_name == 'SITUACAO_NOTA':
                    status_cols_conc[col_name] = col_idx

                if col_name in ['CFOP_XML', 'CFOP_SPED', 'STATUS_CFOP']:
                    cfop_cols_conc[col_name] = col_idx

                if any(substring in col_name for substring in ['VL_', 'ICMS', 'IPI', 'PIS', 'COFINS', 'FCP', 'BC_']):
                    num_format_to_apply = format_currency
                elif col_name == 'CHV_NFE': width = 48
                elif col_name == 'CEST_XML': width = 25
                elif col_name == 'TIPO_NOTA':
                    width = 25
                    status_cols_conc[col_name] = col_idx
                else:
                    try: max_len = max(len(str(col_name)), df_recon_relatorio[col_name].astype(str).map(len).max(), 8) + 2
                    except: max_len = len(str(col_name)) + 5
                    width = min(max_len, 60)

                col_formats_conc[col_idx] = (col_name, width, num_format_to_apply)

            apply_styles_and_rules_v2(ws_conciliacao, df_recon_relatorio, status_cols_conc, cfop_cols_conc, col_formats_conc)
        else:
            logging.warning("DataFrame de conciliação (NF-e, C500, D500) vazio. Aba 'Conciliacao' não será gerada (ou estará vazia).")

        # --- GERAÇÃO DA ABA 'Itens_XML' ---
        if not df_itens_aba.empty:
            logging.info("Gerando aba 'Itens_XML' (motor openpyxl)...")
            df_itens_aba.to_excel(writer, sheet_name='Itens_XML', index=False)
            ws_itens = writer.sheets['Itens_XML']

            status_cols_itens = {}
            cfop_cols_itens = {}
            col_formats_itens = {}

            for col_idx, col_name in enumerate(df_itens_aba.columns):
                num_format_to_apply = None
                width = 18

                if col_name.startswith('STATUS_') or col_name == 'SITUACAO_NOTA':
                    status_cols_itens[col_name] = col_idx

                sped_item_currency_cols = ['VL_OPR_SPED_ITEM', 'VL_BC_ICMS_SPED_ITEM', 'VL_ICMS_SPED_ITEM', 'VL_BC_ICMS_ST_SPED_ITEM', 'VL_ICMS_ST_SPED_ITEM', 'VLR_IPI_SPED_ITEM']

                if (any(substring in col_name for substring in ['VL_', '_SPED', '_CALC', '_XML', 'VLR_', 'DIF_', 'IPI_SPED (Item C170)']) or col_name in sped_item_currency_cols) \
                    and col_name not in ['CFOP_XML', 'CFOP_SPED', 'CFOP_SPED_ITEM', 'CST_ICMS_XML', 'VLR_UNIT', 'CST_ICMS_SPED_ITEM']:
                    num_format_to_apply = format_currency
                    width = 16
                elif col_name == 'pICMS_XML':
                    num_format_to_apply = format_percent
                    width = 10
                elif col_name == 'MVA ORIGINAL':
                    num_format_to_apply = format_mva
                    width = 12
                elif col_name == 'QTD' or col_name == 'VLR_UNIT':
                    num_format_to_apply = format_number
                    width = 14
                elif col_name == 'CHV_NFE': width = 48
                elif col_name == 'CEST': width = 25
                elif col_name == 'TIPO_NOTA':
                    width = 25
                    status_cols_itens[col_name] = col_idx
                elif col_name == 'TIPO_DESTINATARIO': width = 10
                else:
                    try: max_len = max(len(str(col_name)), df_itens_aba[col_name].astype(str).map(len).max(), 8) + 2
                    except: max_len = len(str(col_name)) + 5
                    width = min(max_len, 40)

                col_formats_itens[col_idx] = (col_name, width, num_format_to_apply)

            apply_styles_and_rules_v2(ws_itens, df_itens_aba, status_cols_itens, cfop_cols_itens, col_formats_itens)
        else:
            logging.warning("DataFrame de itens vazio. Aba 'Itens_XML' não será gerada.")

        # --- GERAÇÃO DA ABA 'Aliquota_XML' ---
        if not df_aliquota_aba.empty:
            logging.info("Gerando aba 'Aliquota_XML' (motor openpyxl)...")
            df_aliquota_aba.to_excel(writer, sheet_name='Aliquota_XML', index=False)
            ws_aliquota = writer.sheets['Aliquota_XML']

            col_formats_aliquota = {}
            for col_idx, col_name in enumerate(df_aliquota_aba.columns):
                num_format_to_apply = None
                width = 18

                if col_name == 'Aliquota ICMS (XML)':
                    num_format_to_apply = format_percent
                elif col_name == 'MVA Original (Regra)':
                    num_format_to_apply = format_mva
                    width = 15
                elif col_name in ['VLR_BC_ICMS_XML', 'VLR_ICMS', 'VLR_ICMS_ST', 'VLR_PROD', 'VLR_TOTAL_NF', 'VLR_ICMS_SOMA_SN']:
                    num_format_to_apply = format_currency
                elif col_name == 'CEST': width = 25
                elif col_name == 'TIPO_NOTA': width = 25
                else:
                    try: max_len = max(len(str(col_name)), df_aliquota_aba[col_name].astype(str).map(len).max(), 8) + 2
                    except: max_len = len(str(col_name)) + 5
                    width = min(max_len, 40)

                col_formats_aliquota[col_idx] = (col_name, width, num_format_to_apply)

            apply_styles_and_rules_v2(ws_aliquota, df_aliquota_aba, {}, {}, col_formats_aliquota)

        # --- GERAÇÃO DA ABA 'Totalizadores_Entrada' ---
        if not df_totalizadores_entrada.empty:
            logging.info("Gerando aba 'Totalizadores_Entrada' (motor openpyxl)...")
            df_totalizadores_entrada.to_excel(writer, sheet_name='Totalizadores_Entrada', index=False)
            ws_totalizadores_ent = writer.sheets['Totalizadores_Entrada']

            col_formats_ent = {}
            for col_idx, col_name in enumerate(df_totalizadores_entrada.columns):
                num_format_to_apply = None
                width = 18

                if col_name in ['Total Operação', 'Base de Cálculo ICMS', 'Total ICMS', 'Base de Cálculo ICMS ST', 'Total ICMS ST', 'Total IPI']:
                    num_format_to_apply = format_currency
                    width = 19
                elif col_name == 'Alíquota ICMS':
                    num_format_to_apply = format_aliquota
                    width = 12
                elif col_name == 'Alíquota (SPED)':
                    num_format_to_apply = format_aliquota
                    width = 15
                elif col_name == 'CFOP (SPED)': width = 12
                elif col_name == 'CST (SPED)': width = 10
                elif col_name == 'Descricao CST':
                    num_format_to_apply = None
                    width = 45
                elif col_name == 'QTD Documentos': width = 10

                col_formats_ent[col_idx] = (col_name, width, num_format_to_apply)

            apply_styles_and_rules_v2(ws_totalizadores_ent, df_totalizadores_entrada, {}, {}, col_formats_ent)
        else:
            logging.warning("DataFrame de totalizadores (Entrada) vazio. Aba 'Totalizadores_Entrada' não será gerada.")

        # --- GERAÇÃO DA ABA 'Totalizadores_Saida' ---
        if not df_totalizadores_saida.empty:
            logging.info("Gerando aba 'Totalizadores_Saida' (motor openpyxl)...")
            df_totalizadores_saida.to_excel(writer, sheet_name='Totalizadores_Saida', index=False)
            ws_totalizadores_sai = writer.sheets['Totalizadores_Saida']

            col_formats_sai = {}
            for col_idx, col_name in enumerate(df_totalizadores_saida.columns):
                num_format_to_apply = None
                width = 18

                if col_name in ['Total Operação', 'Base de Cálculo ICMS', 'Total ICMS', 'Base de Cálculo ICMS ST', 'Total ICMS ST', 'Total IPI']:
                    num_format_to_apply = format_currency
                    width = 19
                elif col_name == 'Alíquota ICMS':
                    num_format_to_apply = format_aliquota
                    width = 12
                elif col_name == 'Alíquota (SPED)':
                    num_format_to_apply = format_aliquota
                    width = 15
                elif col_name == 'CFOP (SPED)': width = 12
                elif col_name == 'CST (SPED)': width = 10
                elif col_name == 'Descricao CST':
                    num_format_to_apply = None
                    width = 45
                elif col_name == 'QTD Documentos': width = 10

                col_formats_sai[col_idx] = (col_name, width, num_format_to_apply)

            apply_styles_and_rules_v2(ws_totalizadores_sai, df_totalizadores_saida, {}, {}, col_formats_sai)
        else:
            logging.warning("DataFrame de totalizadores (Saida) vazio. Aba 'Totalizadores_Saida' não será gerada.")

        # --- GERAÇÃO DA ABA 'Dados_CTe_SPED' (D190 Bruto e Limpo) ---
        if not df_cte_bruto_aba.empty:
            logging.info("Gerando aba 'Dados_CTe_SPED' (motor openpyxl)...")
            df_cte_bruto_aba.to_excel(writer, sheet_name='Dados_CTe_SPED', index=False)
            ws_cte = writer.sheets['Dados_CTe_SPED']

            # --- MUDANÇA INICIA AQUI ---
            status_cols_cte = {} # Agora preenchemos este dicionário
            cfop_cols_cte = {}   # Adicionado para consistência
            col_formats_cte = {}

            for col_idx, col_name in enumerate(df_cte_bruto_aba.columns):
                num_format_to_apply = None
                width = 18

                # 1. Identifica colunas de STATUS
                if col_name.startswith('STATUS_') or col_name == 'SITUACAO_CTE':
                    status_cols_cte[col_name] = col_idx # Adiciona ao mapa de status
                    width = 15 # Define uma largura padrão

                # 2. Formata colunas de VALOR (SPED e XML)
                elif col_name in [
                    'VL_OPR_SPED_D190', 'VL_BC_ICMS_SPED_D190', 'VL_ICMS_SPED_D190',
                    'VL_OPR_XML', 'VL_BC_ICMS_XML', 'VL_ICMS_XML'
                ]:
                    num_format_to_apply = format_currency
                    width = 19

                # 3. Formata ALÍQUOTA
                elif col_name == 'ALIQ_ICMS_SPED_D190':
                    num_format_to_apply = format_aliquota
                    width = 12

                # 4. Formata CHAVE
                elif col_name == 'CHV_CTE':
                    width = 48

                # 5. Formata CFOP/CST (SPED e XML)
                elif col_name in [
                    'CST_ICMS_SPED_D190', 'CFOP_SPED_D190',
                    'CFOP_XML', 'CST_XML'
                ]:
                    width = 10
                    if col_name.startswith('CFOP_'):
                         cfop_cols_cte[col_name] = col_idx # Adiciona ao mapa de CFOP
                else:
                    width = 15 # Largura padrão para outras colunas

                col_formats_cte[col_idx] = (col_name, width, num_format_to_apply)

            # 6. Chama a função com os dicionários PREENCHIDOS
            apply_styles_and_rules_v2(
                ws_cte,
                df_cte_bruto_aba,
                status_cols_cte, # Passa o mapa de status
                cfop_cols_cte,   # Passa o mapa de cfop
                col_formats_cte
            )
            # --- FIM DA MUDANÇA ---
        else:
            logging.warning("DataFrame de CT-e (D190) vazio. Aba 'Dados_CTe_SPED' não será gerada.")

        writer.close()

    except Exception as e:
        logging.exception("Ocorreu uma falha crítica na geração do relatório Excel.")
        if writer is not None and hasattr(writer, 'close') and not writer.closed:
            try:
                writer.close()
            except Exception as close_e:
                logging.error(f"Erro ao tentar fechar o ExcelWriter (openpyxl v2) após falha: {close_e}")
        raise