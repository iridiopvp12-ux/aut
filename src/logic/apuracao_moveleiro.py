import logging
import shutil
import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from typing import List, Optional

# ==============================================================================
# 1. FUNÇÕES AUXILIARES E ESTILO
# ==============================================================================

def _limpar_cfop_excel(valor_celula) -> List[str]:
    """Lê uma célula do Excel e retorna lista limpa de CFOPs string"""
    if not valor_celula: return []
    s = str(valor_celula).replace(' ', '').strip()
    partes = s.split('/')
    return [p for p in partes if p.isdigit()]

def _escrever_seguro(ws: Worksheet, linha: int, coluna: int, valor: float):
    """Escreve em células mescladas ou normais."""
    cell = ws.cell(row=linha, column=coluna)
    if isinstance(cell, MergedCell):
        for merged_range in ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                top_left_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                top_left_cell.value = valor
                return cell
    else:
        cell.value = valor
        return cell

def _aplicar_estilo_tabela_sobras(ws, linha, col_inicial, col_final, is_header=False, cor_header="4F81BD"):
    """Aplica bordas e cores para o quadro de sobras."""
    thick = Side(border_style="medium", color="000000")
    thin = Side(border_style="thin", color="000000")
    borda = Border(top=thin, left=thin, right=thin, bottom=thin)

    for c in range(col_inicial, col_final + 1):
        cell = ws.cell(row=linha, column=c)
        cell.border = borda
        if is_header:
            cell.fill = PatternFill(start_color=cor_header, end_color=cor_header, fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            if c == col_final: # Motivo
                cell.alignment = Alignment(horizontal="left")
            elif c in [col_inicial, col_inicial+1]: # CFOP e Aliq
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = '#,##0.00'

def _gerar_relatorio_sobras(ws: Worksheet, df: pd.DataFrame, col_inicio: int, titulo_bloco: str, cor_fundo: str):
    """Gera a tabela lateral com as notas não utilizadas."""
    logging.info(f"[MOVELEIRO] Gerando relatório de sobras: {titulo_bloco}")

    # Filtra o que não foi usado e tem valor relevante
    mask_sobra = (~df['Utilizado']) & (df['Total Operação'] > 0.01)
    df_sobra = df.loc[mask_sobra].copy().sort_values(by='Total Operação', ascending=False)

    if df_sobra.empty:
        return

    # Definição das colunas relativas ao inicio
    C_CFOP = col_inicio
    C_ALIQ = col_inicio + 1
    C_VALOR = col_inicio + 2
    C_BASE = col_inicio + 3
    C_ICMS = col_inicio + 4
    C_MOTIVO = col_inicio + 5

    LINHA = 5 # Linha inicial do relatório

    # Título Geral
    ws.merge_cells(start_row=LINHA-2, start_column=col_inicio, end_row=LINHA-2, end_column=C_MOTIVO)
    cell_title = ws.cell(row=LINHA-2, column=col_inicio, value=f"⚠️ SOBRAS - {titulo_bloco}")
    cell_title.font = Font(bold=True, color="FFFFFF", size=11)
    cell_title.fill = PatternFill(start_color=cor_fundo, end_color=cor_fundo, fill_type="solid")
    cell_title.alignment = Alignment(horizontal="center")

    # Cabeçalhos
    titulos = {C_CFOP: "CFOP", C_ALIQ: "Aliq %", C_VALOR: "Vlr Contábil", C_BASE: "Base Calc", C_ICMS: "ICMS", C_MOTIVO: "Provável Motivo"}
    for col, titulo in titulos.items():
        _escrever_seguro(ws, LINHA-1, col, titulo)
    _aplicar_estilo_tabela_sobras(ws, LINHA-1, col_inicio, C_MOTIVO, is_header=True, cor_header=cor_fundo)

    # Preenchimento
    for _, row in df_sobra.iterrows():
        if LINHA > 200: break # Limite de segurança para não travar excel

        cfop = str(row['CFOP (SPED)']).replace('.0', '')
        aliq = row['Alíquota (SPED)']

        # Lógica simples de "Motivo"
        motivo = "Não mapeado"
        if aliq == 0: motivo = "Alíquota Zero"
        elif cfop in ['1403', '2403', '6403', '5403']: motivo = "Subst. Tributária"
        elif cfop.startswith('59') or cfop.startswith('69') or cfop.startswith('19') or cfop.startswith('29'): motivo = "Outras/Isentas"

        _escrever_seguro(ws, LINHA, C_CFOP, cfop)
        _escrever_seguro(ws, LINHA, C_ALIQ, aliq)
        _escrever_seguro(ws, LINHA, C_VALOR, row['Total Operação'])
        _escrever_seguro(ws, LINHA, C_BASE, row['Base de Cálculo ICMS'])
        _escrever_seguro(ws, LINHA, C_ICMS, row['Total ICMS'])
        _escrever_seguro(ws, LINHA, C_MOTIVO, motivo)

        _aplicar_estilo_tabela_sobras(ws, LINHA, col_inicio, C_MOTIVO, is_header=False)
        LINHA += 1


def _escrever_caixa_informativa_difal(ws: Worksheet, df_difal: pd.DataFrame, row_start=15, col_start=14):
    """Cria uma caixa visual (aviso) mostrando quais CFOPs tiveram abatimento de DIFAL."""
    if df_difal is None or df_difal.empty: return

    thick = Side(border_style="medium", color="000000")
    thin = Side(border_style="thin", color="000000")
    border_box = Border(top=thick, left=thick, right=thick, bottom=thick)
    border_row = Border(left=thick, right=thick, bottom=thin)

    fill_header = PatternFill(start_color="C0504D", end_color="C0504D", fill_type="solid")
    font_header = Font(bold=True, color="FFFFFF")

    ws.merge_cells(start_row=row_start, start_column=col_start, end_row=row_start, end_column=col_start+1)
    cell_header = ws.cell(row=row_start, column=col_start, value="⚠️ ABATIMENTO DIFAL (C101)")
    cell_header.fill = fill_header
    cell_header.font = font_header
    cell_header.alignment = Alignment(horizontal='center')
    cell_header.border = border_box

    r = row_start + 1
    ws.cell(row=r, column=col_start, value="CFOP").font = Font(bold=True)
    ws.cell(row=r, column=col_start+1, value="Base Abatida").font = Font(bold=True)

    r += 1
    total_abatido = 0.0
    for _, row in df_difal.iterrows():
        cfop = str(row['CFOP'])
        valor = row['VALOR_BASE_DIFAL']
        total_abatido += valor

        c1 = ws.cell(row=r, column=col_start, value=cfop)
        c2 = ws.cell(row=r, column=col_start+1, value=valor)
        c1.alignment = Alignment(horizontal='center')
        c2.number_format = '#,##0.00'
        c1.border = border_row
        c2.border = border_row
        r += 1

    ws.cell(row=r, column=col_start, value="TOTAL:").font = Font(bold=True)
    c_total = ws.cell(row=r, column=col_start+1, value=total_abatido)
    c_total.font = Font(bold=True)
    c_total.number_format = '#,##0.00'
    c_total.border = Border(top=thick, left=thick, right=thick, bottom=thick)

# ==============================================================================
# 2. LÓGICA ESPECÍFICA: SETOR MOVELEIRO (ENTRADAS)
# ==============================================================================

def preencher_quadro_entradas_moveleiro(ws: Worksheet, df_totalizadores: pd.DataFrame):
    logging.info("[MOVELEIRO] Iniciando preenchimento ENTRADAS...")
    if df_totalizadores is None or df_totalizadores.empty: return

    df = df_totalizadores.copy()
    df['Utilizado'] = False # Inicializa rastreamento

    cols_num = ['Alíquota (SPED)', 'Alíquota ICMS', 'Base de Cálculo ICMS', 'Total ICMS', 'Total Operação']
    for col in cols_num:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0.0)
        else:
            df[col] = 0.0

    # --- REGRA 1: Linhas 17 a 36 ---
    for linha in range(17, 37):
        cfop_cell = ws.cell(row=linha, column=1).value
        lista_cfops = _limpar_cfop_excel(cfop_cell)
        if not lista_cfops: continue

        mask = (df['CFOP (SPED)'].isin(lista_cfops)) & (df['Alíquota ICMS'] > 7.0)

        # Marca como utilizado se encontrou algo
        if mask.any(): df.loc[mask, 'Utilizado'] = True

        df_filtered = df.loc[mask]
        if not df_filtered.empty:
            soma_base = df_filtered['Base de Cálculo ICMS'].sum()
            soma_icms = df_filtered['Total ICMS'].sum()
            if soma_base > 0: _escrever_seguro(ws, linha, 2, soma_base)
            if soma_icms > 0: _escrever_seguro(ws, linha, 3, soma_icms)

    # --- REGRA 2: Linhas 42 a 45 ---
    for linha in range(42, 46):
        cfop_cell = ws.cell(row=linha, column=1).value
        lista_cfops = _limpar_cfop_excel(cfop_cell)
        if not lista_cfops: continue

        mask = (df['CFOP (SPED)'].isin(lista_cfops)) & (np.isclose(df['Alíquota (SPED)'], 12.0, atol=0.1))

        if mask.any(): df.loc[mask, 'Utilizado'] = True

        df_filtered = df.loc[mask]
        if not df_filtered.empty:
            soma_base = df_filtered['Base de Cálculo ICMS'].sum()
            soma_icms = df_filtered['Total ICMS'].sum()
            if soma_base > 0: _escrever_seguro(ws, linha, 2, soma_base)
            if soma_icms > 0: _escrever_seguro(ws, linha, 3, soma_icms)

    # --- TOTALIZADOR CRÉDITO ---
    total_credito = df['Total ICMS'].sum()
    if total_credito > 0:
        _escrever_seguro(ws, 72, 5, total_credito)

    # --- RELATÓRIO SOBRAS ENTRADAS (COLUNA R / 18) ---
    _gerar_relatorio_sobras(ws, df, 18, "ENTRADAS", "305496")

# ==============================================================================
# 3. LÓGICA ESPECÍFICA: SETOR MOVELEIRO (SAÍDAS)
# ==============================================================================

def preencher_quadro_saidas_moveleiro(ws: Worksheet, df_saidas: pd.DataFrame, df_base_difal: pd.DataFrame = None):
    logging.info("[MOVELEIRO] Iniciando preenchimento SAÍDAS...")
    if df_saidas is None or df_saidas.empty: return

    df = df_saidas.copy()
    df['Utilizado'] = False # Inicializa rastreamento

    # Prepara DF de DIFAL
    mapa_difal = {}
    if df_base_difal is not None and not df_base_difal.empty:
        df_base_difal['CFOP'] = df_base_difal['CFOP'].astype(str).str.strip()
        mapa_difal = df_base_difal.set_index('CFOP')['VALOR_BASE_DIFAL'].to_dict()
        try:
            # Desenha a caixa de DIFAL na coluna N (14)
            _escrever_caixa_informativa_difal(ws, df_base_difal, row_start=3, col_start=14)
        except Exception as e:
            logging.warning(f"Não foi possível desenhar caixa de DIFAL: {e}")

    cols_num = ['Alíquota (SPED)', 'Base de Cálculo ICMS', 'Total ICMS', 'Total Operação']
    for col in cols_num:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0.0)
        else: df[col] = 0.0

    # --- REGRA 3: Linhas 3 a 15 (Aliq == 12%) ---
    for linha in range(3, 16):
        cfop_cell = ws.cell(row=linha, column=9).value
        lista_cfops = _limpar_cfop_excel(cfop_cell)
        if not lista_cfops: continue

        mask = (df['CFOP (SPED)'].isin(lista_cfops)) & \
               (np.isclose(df['Alíquota (SPED)'], 12.0, atol=0.1))

        if mask.any(): df.loc[mask, 'Utilizado'] = True

        df_filtered = df.loc[mask]
        if not df_filtered.empty:
            soma_base = df_filtered['Base de Cálculo ICMS'].sum()

            abatimento_difal = 0.0
            for cfop_lista in lista_cfops:
                if cfop_lista in mapa_difal:
                    abatimento_difal += mapa_difal[cfop_lista]

            valor_final_base = soma_base - abatimento_difal
            if valor_final_base > 0:
                _escrever_seguro(ws, linha, 10, valor_final_base)

    # --- TOTALIZADOR DÉBITO ---
    total_debito = df['Total ICMS'].sum()
    if total_debito > 0:
        _escrever_seguro(ws, 61, 5, total_debito)

    # --- RELATÓRIO SOBRAS SAÍDAS (COLUNA Y / 25) ---
    # Colocado na coluna 25 (Y) para ficar longe da caixa de DIFAL (N/14) e do rel. de Entradas (R/18 a W/23)
    _gerar_relatorio_sobras(ws, df, 25, "SAÍDAS", "C65911")

# ==============================================================================
# 4. FUNÇÃO PRINCIPAL (ORQUESTRADOR)
# ==============================================================================

def preencher_template_moveleiro(template_path: Path, df_entradas: pd.DataFrame, df_saidas: pd.DataFrame, df_base_difal: pd.DataFrame = None) -> str:
    logging.info(f"[MOVELEIRO] Processando arquivo base: {template_path}")

    if (df_entradas is None or df_entradas.empty) and (df_saidas is None or df_saidas.empty):
        return str(template_path)

    try:
        path_origem = Path(template_path)
        pasta = path_origem.parent
        novo_nome = f"{path_origem.stem}_MOVELEIRO_PREENCHIDA{path_origem.suffix}"
        path_destino = pasta / novo_nome

        shutil.copy(path_origem, path_destino)
        wb = load_workbook(path_destino)
        ws = wb["Entradas"] if "Entradas" in wb.sheetnames else wb.active

        if df_entradas is not None and not df_entradas.empty:
            preencher_quadro_entradas_moveleiro(ws, df_entradas)

        if df_saidas is not None and not df_saidas.empty:
            preencher_quadro_saidas_moveleiro(ws, df_saidas, df_base_difal)

        wb.save(path_destino)
        logging.info(f"[MOVELEIRO] Sucesso! Arquivo gerado: {path_destino}")
        return str(path_destino)

    except Exception as e:
        logging.error(f"[MOVELEIRO] Erro fatal ao gerar apuração: {e}")
        raise e