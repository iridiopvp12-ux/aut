import logging
import shutil
import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from typing import List, Optional

# ==============================================================================
# 1. FUNÇÕES AUXILIARES E ESTILO
# ==============================================================================

def _limpar_cfop_excel(valor_celula) -> List[str]:
    if not valor_celula: return []
    s = str(valor_celula).replace(' ', '').strip()
    partes = s.split('/')
    return [p for p in partes if p.isdigit()]

def _normalizar_aliquota(valor_celula) -> float:
    if valor_celula is None: return 0.0
    try:
        val = float(valor_celula)
        if 0 < val < 1: return round(val * 100, 2)
        return val
    except (ValueError, TypeError): return 0.0

def _escrever_seguro(ws: Worksheet, linha: int, coluna: int, valor: float):
    cell = ws.cell(row=linha, column=coluna)
    if isinstance(cell, MergedCell):
        for merged_range in ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                top_left_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                top_left_cell.value = valor
                return cell
    else:
        cell.value = valor
        return cell

def _aplicar_estilo_tabela(ws, linha, col_inicial, col_final, is_header=False, cor_header="4F81BD"):
    thin = Side(border_style="thin", color="000000")
    borda = Border(top=thin, left=thin, right=thin, bottom=thin)
    for c in range(col_inicial, col_final + 1):
        cell = ws.cell(row=linha, column=c)
        cell.border = borda
        if is_header:
            cell.fill = PatternFill(start_color=cor_header, end_color=cor_header, fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            if c in [col_inicial, col_inicial+1]: cell.alignment = Alignment(horizontal="center")
            elif c == col_final: cell.alignment = Alignment(horizontal="left")
            else:
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = '#,##0.00'

def _escrever_placar_geral(ws, df, col_inicio, titulo_bloco, cor_fundo):
    total_contabil = df['Total Operação'].sum()
    total_base = df['Base de Cálculo ICMS'].sum()
    total_icms = df['Total ICMS'].sum()

    ws.cell(row=1, column=col_inicio).value = f"TOTAL GERAL SPED ({titulo_bloco})"
    ws.merge_cells(start_row=1, start_column=col_inicio, end_row=1, end_column=col_inicio+2)

    cell_title = ws.cell(row=1, column=col_inicio)
    cell_title.fill = PatternFill(start_color=cor_fundo, end_color=cor_fundo, fill_type="solid")
    cell_title.font = Font(bold=True, color="FFFFFF", size=11)
    cell_title.alignment = Alignment(horizontal="center")

    headers = ["Vlr Contábil", "Base Calc", "Vlr ICMS"]
    for i, h in enumerate(headers):
        c = ws.cell(row=2, column=col_inicio + i)
        c.value = h
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")
        c.border = Border(bottom=Side(style='thin'))

    vals = [total_contabil, total_base, total_icms]
    for i, v in enumerate(vals):
        c = ws.cell(row=3, column=col_inicio + i)
        c.value = v
        c.number_format = '#,##0.00'
        c.font = Font(bold=True, size=11)
        c.alignment = Alignment(horizontal="right")

# ==============================================================================
# 2. ENTRADAS (06-26, 28-52, 53-56)
# ==============================================================================

def preencher_quadro_entradas(ws: Worksheet, df_totalizadores: pd.DataFrame):
    logging.info("Iniciando preenchimento ENTRADAS...")
    df = df_totalizadores.copy()
    df['Utilizado'] = False

    cols = ['Alíquota (SPED)', 'Alíquota ICMS', 'Total Operação', 'Base de Cálculo ICMS', 'Total ICMS']
    for col in cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0.0)

    _escrever_placar_geral(ws, df, col_inicio=17, titulo_bloco="ENTRADAS", cor_fundo="203764")

    # --- Lógica ---
    def processar_linha_padrao(linha_num, regra_tipo):
        cfop_cell = ws.cell(row=linha_num, column=2).value
        aliq_cell = ws.cell(row=linha_num, column=6).value
        lista_cfops = _limpar_cfop_excel(cfop_cell)
        aliq_alvo = _normalizar_aliquota(aliq_cell)
        if not lista_cfops or aliq_cell is None: return

        # Filtro 1: Compatibilidade básica (Tolerância 0.5)
        mask_target = (df['CFOP (SPED)'].isin(lista_cfops) &
                       np.isclose(df['Alíquota (SPED)'], aliq_alvo, atol=0.5))

        # --- PROTEÇÃO CONTRA SIMPLES NACIONAL ---
        # Se a linha da planilha pede 4% ou mais, NÃO aceitar notas < 4.0 (Simples)
        # Isso evita que a tolerância de 0.5 puxe notas de 3.5% para a linha de 4.0%
        if aliq_alvo >= 4.0:
            mask_target = mask_target & (df['Alíquota (SPED)'] >= 4.0)
        # ----------------------------------------

        if not mask_target.any(): return

        # Filtro 2: Validação de Igualdade Padrão (-0.5% a +0.01%)
        condicao_igual_padrao = (
            (df['Alíquota ICMS'] >= (df['Alíquota (SPED)'] - 0.5)) &
            (df['Alíquota ICMS'] <= (df['Alíquota (SPED)'] + 0.01))
        )

        if regra_tipo == 'IGUAL':
            # --- REGRA DE EXCEÇÃO (CFOP 2102/2910 e Aliq > 7) ---
            condicao_excecao_cfop = (
                df['CFOP (SPED)'].isin(['2102', '2910']) &
                (df['Alíquota (SPED)'] > 7.0)
            )
            mask = mask_target & (condicao_igual_padrao | condicao_excecao_cfop)

        else: # DIFERENTE
            if aliq_alvo <= 7.0:
                mask = mask_target # Aceita tudo se Alíquota da Planilha for <= 7%
            else:
                mask = mask_target & (~condicao_igual_padrao)

        if mask.any(): df.loc[mask, 'Utilizado'] = True

        df_filtered = df.loc[mask]
        if df_filtered.empty: return

        soma_contabil = df_filtered['Total Operação'].sum()
        soma_base = df_filtered['Base de Cálculo ICMS'].sum()
        soma_icms = df_filtered['Total ICMS'].sum()
        if soma_contabil > 0: _escrever_seguro(ws, linha_num, 3, soma_contabil)
        if soma_base > 0: _escrever_seguro(ws, linha_num, 5, soma_base)
        if soma_icms > 0: _escrever_seguro(ws, linha_num, 13, soma_icms)

    def processar_simples_entradas(linha_num):
        cfop_cell = ws.cell(row=linha_num, column=2).value
        lista_cfops = _limpar_cfop_excel(cfop_cell)
        if not lista_cfops: return
        mask = (df['CFOP (SPED)'].isin(lista_cfops) & (df['Alíquota (SPED)'] < 4.0))
        if mask.any(): df.loc[mask, 'Utilizado'] = True
        df_filtered = df.loc[mask]
        if df_filtered.empty: return

        soma_contabil = df_filtered['Total Operação'].sum()
        soma_base = df_filtered['Base de Cálculo ICMS'].sum()
        soma_icms = df_filtered['Total ICMS'].sum()
        if soma_contabil > 0: _escrever_seguro(ws, linha_num, 3, soma_contabil)
        if soma_base > 0: _escrever_seguro(ws, linha_num, 5, soma_base)
        if soma_icms > 0: _escrever_seguro(ws, linha_num, 7, soma_icms)

    def listar_sobras_entradas():
        logging.info("Listando ENTRADAS não processadas...")
        mask_sobra = (~df['Utilizado']) & (df['Total Operação'] > 0.01)
        df_sobra = df.loc[mask_sobra].copy().sort_values(by='Total Operação', ascending=False)

        C_CFOP, C_ALIQ, C_VALOR, C_BASE, C_ICMS, C_MOTIVO = 15, 16, 17, 18, 19, 20
        LINHA = 5

        titulos = {C_CFOP: "CFOP", C_ALIQ: "Aliq %", C_VALOR: "Vlr Contábil", C_BASE: "Base Calc", C_ICMS: "ICMS", C_MOTIVO: "Motivo (Entrada)"}
        for col, titulo in titulos.items(): _escrever_seguro(ws, 4, col, titulo)
        _aplicar_estilo_tabela(ws, 4, 15, 20, is_header=True, cor_header="305496")

        for _, row in df_sobra.iterrows():
            if LINHA > 100: break
            cfop = str(row['CFOP (SPED)']).replace('.0', '')
            aliq = row['Alíquota (SPED)']
            motivo = "Não mapeado"
            if aliq == 0: motivo = "Alíquota Zero"
            if cfop in ['1403', '2403']: motivo = "ST (Aliq 0)"

            _escrever_seguro(ws, LINHA, C_CFOP, cfop)
            _escrever_seguro(ws, LINHA, C_ALIQ, aliq)
            _escrever_seguro(ws, LINHA, C_VALOR, row['Total Operação'])
            _escrever_seguro(ws, LINHA, C_BASE, row['Base de Cálculo ICMS'])
            _escrever_seguro(ws, LINHA, C_ICMS, row['Total ICMS'])
            _escrever_seguro(ws, LINHA, C_MOTIVO, motivo)
            _aplicar_estilo_tabela(ws, LINHA, 15, 20, is_header=False)
            LINHA += 1

    # --- EXECUÇÃO ENTRADAS (06-26, 28-52, 53-56) ---
    for linha in range(6, 27): processar_linha_padrao(linha, 'IGUAL')
    for linha in range(28, 53): processar_linha_padrao(linha, 'DIFERENTE')
    for linha in range(53, 57): processar_simples_entradas(linha)
    listar_sobras_entradas()

# ==============================================================================
# 3. SAÍDAS (75-87, 98-114, 116-148)
# ==============================================================================

def preencher_quadro_saidas(ws: Worksheet, df_saidas: pd.DataFrame):
    logging.info("Iniciando preenchimento SAÍDAS...")
    if df_saidas is None or df_saidas.empty: return
    df = df_saidas.copy()
    df['Utilizado'] = False

    cols = ['Alíquota (SPED)', 'Alíquota ICMS', 'Total Operação', 'Base de Cálculo ICMS', 'Total ICMS']
    for col in cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0.0)

    _escrever_placar_geral(ws, df, col_inicio=24, titulo_bloco="SAÍDAS", cor_fundo="974706")

    COL_CONTABIL, COL_BASE, COL_ICMS = 5, 7, 13

    def _checar_regime_base(ws: Worksheet, linha: int) -> str:
        valor = ws.cell(row=linha, column=14).value
        if valor:
            valor = str(valor).strip().upper()
            if 'BASE CHEIA' in valor: return 'CHEIA'
            if 'BASE REDUZIDA' in valor: return 'REDUZIDA'
        return 'NORMAL'

    def processar_saida_padrao(linha_num, regra_tipo, regime_base='NORMAL'):
        cfop_cell = ws.cell(row=linha_num, column=2).value
        aliq_cell = ws.cell(row=linha_num, column=8).value
        lista_cfops = _limpar_cfop_excel(cfop_cell)
        aliq_alvo = _normalizar_aliquota(aliq_cell)
        if not lista_cfops: return

        # Filtro 1 (Tolerância AJUSTADA para 0.5)
        mask_target = (df['CFOP (SPED)'].isin(lista_cfops))
        if aliq_cell is not None:
            mask_target = mask_target & np.isclose(df['Alíquota (SPED)'], aliq_alvo, atol=0.5)

        # --- PROTEÇÃO CONTRA SIMPLES NACIONAL (Também nas Saídas) ---
        if aliq_cell is not None and aliq_alvo >= 4.0:
            mask_target = mask_target & (df['Alíquota (SPED)'] >= 4.0)
        # ------------------------------------------------------------

        if not mask_target.any(): return

        # Filtro 2 (Assimétrico -0.5%)
        condicao_igual = (
            (df['Alíquota ICMS'] >= (df['Alíquota (SPED)'] - 0.5)) &
            (df['Alíquota ICMS'] <= (df['Alíquota (SPED)'] + 0.01))
        )

        if regra_tipo == 'IGUAL':
            mask = mask_target & condicao_igual
        elif regra_tipo == 'DIFERENTE':
            mask = mask_target & (~condicao_igual)
        elif regra_tipo == 'GENERICA':
            mask = mask_target

        if regime_base == 'CHEIA':
            mask = mask_target & condicao_igual
        elif regime_base == 'REDUZIDA':
            mask = mask_target & (~condicao_igual)

        if mask.any(): df.loc[mask, 'Utilizado'] = True
        df_filtered = df.loc[mask]
        if df_filtered.empty: return

        soma_contabil = df_filtered['Total Operação'].sum()
        soma_base = df_filtered['Base de Cálculo ICMS'].sum()
        soma_icms = df_filtered['Total ICMS'].sum()

        if soma_contabil > 0: _escrever_seguro(ws, linha_num, COL_CONTABIL, soma_contabil)
        if soma_base > 0: _escrever_seguro(ws, linha_num, COL_BASE, soma_base)
        if soma_icms > 0: _escrever_seguro(ws, linha_num, COL_ICMS, soma_icms)

    def processar_saida_simples(linha_num):
        cfop_cell = ws.cell(row=linha_num, column=2).value
        lista_cfops = _limpar_cfop_excel(cfop_cell)
        if not lista_cfops: return
        mask = (df['CFOP (SPED)'].isin(lista_cfops) & (df['Alíquota (SPED)'] < 4.0))
        if mask.any(): df.loc[mask, 'Utilizado'] = True
        df_filtered = df.loc[mask]
        if df_filtered.empty: return
        soma_contabil = df_filtered['Total Operação'].sum()
        soma_base = df_filtered['Base de Cálculo ICMS'].sum()
        soma_icms = df_filtered['Total ICMS'].sum()
        if soma_contabil > 0: _escrever_seguro(ws, linha_num, COL_CONTABIL, soma_contabil)
        if soma_base > 0: _escrever_seguro(ws, linha_num, COL_BASE, soma_base)
        if soma_icms > 0: _escrever_seguro(ws, linha_num, COL_ICMS, soma_icms)

    def listar_sobras_saidas():
        logging.info("Listando SAÍDAS não processadas...")
        mask_sobra = (~df['Utilizado']) & (df['Total Operação'] > 0.01)
        df_sobra = df.loc[mask_sobra].copy().sort_values(by='Total Operação', ascending=False)

        C_CFOP, C_ALIQ, C_VALOR, C_BASE, C_ICMS, C_MOTIVO = 22, 23, 24, 25, 26, 27
        LINHA = 5

        titulos = {C_CFOP: "CFOP", C_ALIQ: "Aliq %", C_VALOR: "Vlr Contábil", C_BASE: "Base Calc", C_ICMS: "ICMS", C_MOTIVO: "Motivo (Saída)"}
        for col, titulo in titulos.items(): _escrever_seguro(ws, 4, col, titulo)
        _aplicar_estilo_tabela(ws, 4, 22, 27, is_header=True, cor_header="C65911")

        for _, row in df_sobra.iterrows():
            if LINHA > 100: break
            cfop = str(row['CFOP (SPED)']).replace('.0', '')
            aliq = row['Alíquota (SPED)']
            motivo = "Não mapeado"
            if aliq == 0: motivo = "Alíquota Zero"

            _escrever_seguro(ws, LINHA, C_CFOP, cfop)
            _escrever_seguro(ws, LINHA, C_ALIQ, aliq)
            _escrever_seguro(ws, LINHA, C_VALOR, row['Total Operação'])
            _escrever_seguro(ws, LINHA, C_BASE, row['Base de Cálculo ICMS'])
            _escrever_seguro(ws, LINHA, C_ICMS, row['Total ICMS'])
            _escrever_seguro(ws, LINHA, C_MOTIVO, motivo)
            _aplicar_estilo_tabela(ws, LINHA, 22, 27, is_header=False)
            LINHA += 1

    # --- EXECUÇÃO SAÍDAS (75-87, 98-114, 116-148) ---
    for linha in range(75, 88):   processar_saida_padrao(linha, 'DIFERENTE')
    for linha in range(98, 115):  processar_saida_padrao(linha, 'IGUAL')

    for linha in range(116, 149): # Bloco 3
        # Exceção Coluna N (116 a 121)
        if 116 <= linha <= 121:
            regime = _checar_regime_base(ws, linha)
            if regime != 'NORMAL': processar_saida_padrao(linha, 'GENERICA', regime_base=regime)
            else: processar_saida_padrao(linha, 'GENERICA')

        # Exceção Simples (122 e 130)
        elif linha == 122 or linha == 130:
            processar_saida_simples(linha)

        # Resto (Genérica)
        else:
            processar_saida_padrao(linha, 'GENERICA')

    listar_sobras_saidas()

# ==============================================================================
# 4. PRINCIPAL
# ==============================================================================

def preencher_template_apuracao(template_path: Path, df_entradas: pd.DataFrame, df_saidas: pd.DataFrame = None) -> str:
    logging.info(f"Processando arquivo base: {template_path}")
    if (df_entradas is None or df_entradas.empty) and (df_saidas is None or df_saidas.empty):
        return str(template_path)

    try:
        path_origem = Path(template_path)
        pasta = path_origem.parent
        novo_nome = f"{path_origem.stem}_PREENCHIDA{path_origem.suffix}"
        path_destino = pasta / novo_nome

        shutil.copy(path_origem, path_destino)
        wb = load_workbook(path_destino)
        ws = wb["Entradas"] if "Entradas" in wb.sheetnames else wb.active

        if df_entradas is not None and not df_entradas.empty:
            preencher_quadro_entradas(ws, df_entradas)
        if df_saidas is not None and not df_saidas.empty:
            preencher_quadro_saidas(ws, df_saidas)

        wb.save(path_destino)
        logging.info(f"Sucesso! Arquivo gerado: {path_destino}")
        return str(path_destino)

    except Exception as e:
        logging.error(f"Erro fatal: {e}")
        raise e
           #  Retorna o caminho do arquivo preenchido