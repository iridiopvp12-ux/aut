# app/fiscal/template_generator.py
import json
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment
from typing import List, Dict, Any
import logging

def gerar_template_de_regras(regras_path: str, output_path: str):
    """
    Lê o arquivo de regras JSON e gera um template Excel (.xlsx)
    com todos os labels organizados por tipo de cálculo.
    """
    try:
        with open(regras_path, 'r', encoding='utf-8') as f:
            regras = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError) as e:
        logging.error(f"Erro ao carregar 'regras_apuracao.json': {e}")
        raise

    # Agrupa as regras por tipo
    regras_agrupadas: Dict[str, List[Dict[str, Any]]] = {}
    for regra in regras:
        tipo = regra.get('tipo', 'sem_tipo')
        if tipo not in regras_agrupadas:
            regras_agrupadas[tipo] = []
        regras_agrupadas[tipo].append(regra)

    # Cria o Workbook e a planilha
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Apuracao"

    # Define estilos
    font_titulo = Font(bold=True, size=14)
    font_cabecalho = Font(bold=True, size=12)
    alignment_center = Alignment(horizontal='center')

    ws['A1'] = "Template de Apuração Baseado nas Regras"
    ws['A1'].font = font_titulo
    ws['A1'].alignment = alignment_center
    ws.merge_cells('A1:C1')

    current_row = 3

    # Mapeia os tipos para títulos mais amigáveis
    titulos_grupos = {
        'soma_df': "1. Valores Base (Calculados a partir do SPED/XML)",
        'soma_celulas': "2. Subtotais (Soma de outros campos)",
        'formula': "3. Fórmulas (Cálculos finais)"
    }

    # Itera sobre os grupos na ordem desejada e escreve no Excel
    for tipo in ['soma_df', 'soma_celulas', 'formula']:
        if tipo in regras_agrupadas:
            # Escreve o título do grupo
            ws[f'A{current_row}'] = titulos_grupos.get(tipo, f"Grupo: {tipo}")
            ws[f'A{current_row}'].font = font_cabecalho
            ws.merge_cells(f'A{current_row}:C{current_row}')
            current_row += 1

            # Escreve os labels das regras do grupo
            for regra in regras_agrupadas[tipo]:
                label = regra.get('label')
                if label:
                    ws[f'A{current_row}'] = label
                current_row += 1
            current_row += 1 # Adiciona uma linha em branco entre os grupos

    # Ajusta a largura da coluna A
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 20

    try:
        wb.save(output_path)
        logging.info(f"Template de apuração gerado com sucesso em: {output_path}")
    except Exception as e:
        logging.error(f"Falha ao salvar o template em '{output_path}': {e}")
        raise
