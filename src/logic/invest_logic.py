import sys
import os
import logging
import pandas as pd
import xml.etree.ElementTree as ET
from pathlib import Path
from datetime import datetime
from collections import Counter
from typing import Optional, Callable

# --- IMPORTAÇÕES PARA ESTILO EXCEL ---
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# -----------------------------
# 1. PARSER XML PADRÃO (ATUALIZADO COM PIS/COFINS)
# -----------------------------
def ler_xmls_diretamente(pasta_xml: Path, progress_callback: Optional[Callable[[int, int], None]] = None) -> pd.DataFrame:
    dados = []
    arquivos = list(pasta_xml.glob('*.xml'))
    total_arquivos = len(arquivos)

    if total_arquivos == 0:
        return pd.DataFrame()

    ns = {'nfe': 'http://www.portalfiscal.inf.br/nfe'}

    for i, arquivo in enumerate(arquivos):
        if progress_callback:
            progress_callback(i + 1, total_arquivos)

        try:
            tree = ET.parse(arquivo)
            root = tree.getroot()
            infNFe = root.find('.//nfe:infNFe', ns)
            if infNFe is None:
                infNFe = root.find('.//infNFe')
                if infNFe is None: continue

            def get_text(node, tag):
                if node is None: return ''
                el = node.find(f'nfe:{tag}', ns)
                if el is None: el = node.find(tag)
                return el.text if el is not None else ''

            def get_float(node, tag):
                val = get_text(node, tag)
                if val:
                    return float(val.replace(',', '.'))
                return 0.0

            # --- CABEÇALHO ---
            ide = infNFe.find('.//nfe:ide', ns) or infNFe.find('ide')
            dest = infNFe.find('.//nfe:dest', ns) or infNFe.find('dest')
            prot = root.find('.//nfe:protNFe', ns) or root.find('.//protNFe')

            nNF = get_text(ide, 'nNF')
            dhEmi = get_text(ide, 'dhEmi')[:10]
            cnpj_dest = get_text(dest, 'CNPJ')

            uf_dest = ''
            if dest is not None:
                ender = dest.find('.//nfe:enderDest', ns) or dest.find('enderDest')
                if ender is not None:
                    uf_dest = get_text(ender, 'UF')

            protocolo = ''
            if prot:
                infProt = prot.find('.//nfe:infProt', ns) or prot.find('infProt')
                if infProt is not None:
                    protocolo = get_text(infProt, 'nProt')

            # --- ITENS ---
            dets = infNFe.findall('.//nfe:det', ns) or infNFe.findall('det')

            for det in dets:
                prod = det.find('.//nfe:prod', ns) or det.find('prod')
                imposto = det.find('.//nfe:imposto', ns) or det.find('imposto')

                if prod is None: continue

                cProd = get_text(prod, 'cProd')
                xProd = get_text(prod, 'xProd')
                NCM = get_text(prod, 'NCM')
                CFOP = get_text(prod, 'CFOP')

                vProd = get_float(prod, 'vProd')
                qCom = get_float(prod, 'qCom')
                vUnCom = get_float(prod, 'vUnCom')
                vFrete = get_float(prod, 'vFrete')
                vSeg = get_float(prod, 'vSeg')
                vDesc = get_float(prod, 'vDesc')
                vOutro = get_float(prod, 'vOutro')

                cst = ''; vBC = 0.0; pICMS = 0.0; vICMS = 0.0
                vIPI = 0.0; vIPIDevol = 0.0; vICMSST = 0.0; vFCPST = 0.0
                pCredSN = 0.0; vCredICMSSN = 0.0; vICMSUFDest = 0.0

                # Variáveis PIS/COFINS
                cst_pis = ''; vBC_pis = 0.0; pPIS = 0.0; vPIS = 0.0
                cst_cofins = ''; vBC_cofins = 0.0; pCOFINS = 0.0; vCOFINS = 0.0

                if imposto is not None:
                    # ICMS
                    icms_node = imposto.find('.//nfe:ICMS', ns) or imposto.find('ICMS')
                    if icms_node is not None:
                        for child in icms_node:
                            cst_val = get_text(child, 'CST') or get_text(child, 'CSOSN')
                            if cst_val: cst = cst_val
                            vBC = get_float(child, 'vBC')
                            pICMS = get_float(child, 'pICMS')
                            vICMS = get_float(child, 'vICMS')
                            vICMSST = get_float(child, 'vICMSST')
                            vFCPST = get_float(child, 'vFCPST')
                            pCredSN = get_float(child, 'pCredSN')
                            vCredICMSSN = get_float(child, 'vCredICMSSN')
                            if cst: break

                    # IPI
                    ipi_node = imposto.find('.//nfe:IPI', ns) or imposto.find('IPI')
                    if ipi_node is not None:
                        ipi_trib = ipi_node.find('.//nfe:IPITrib', ns) or ipi_node.find('IPITrib')
                        if ipi_trib is not None:
                            vIPI = get_float(ipi_trib, 'vIPI')
                        else:
                            vIPI = get_float(ipi_node, 'vIPI')

                    # IPI Devol
                    impostoDevol = det.find('.//nfe:impostoDevol', ns) or det.find('impostoDevol')
                    if impostoDevol is not None:
                        vIPIDevol = get_float(impostoDevol, 'vIPIDevol')

                    # DIFAL
                    icms_uf = imposto.find('.//nfe:ICMSUFDest', ns) or imposto.find('ICMSUFDest')
                    if icms_uf is not None:
                        vICMSUFDest = get_float(icms_uf, 'vICMSUFDest')

                    # PIS
                    pis_node = imposto.find('.//nfe:PIS', ns) or imposto.find('PIS')
                    if pis_node is not None:
                        for child in pis_node:
                            cst_pis_val = get_text(child, 'CST')
                            if cst_pis_val: cst_pis = cst_pis_val
                            vBC_pis = get_float(child, 'vBC')
                            pPIS = get_float(child, 'pPIS')
                            vPIS = get_float(child, 'vPIS')

                    # COFINS
                    cofins_node = imposto.find('.//nfe:COFINS', ns) or imposto.find('COFINS')
                    if cofins_node is not None:
                        for child in cofins_node:
                            cst_cofins_val = get_text(child, 'CST')
                            if cst_cofins_val: cst_cofins = cst_cofins_val
                            vBC_cofins = get_float(child, 'vBC')
                            pCOFINS = get_float(child, 'pCOFINS')
                            vCOFINS = get_float(child, 'vCOFINS')

                vItemContabil = (vProd + vIPI + vICMSST + vFrete + vSeg + vOutro + vFCPST) - vDesc

                dados.append({
                    'n da nf': nNF, 'cnpj': cnpj_dest, 'uf': uf_dest, 'data': dhEmi, 'cst': cst,
                    'qnt': qCom, 'vl unit': vUnCom, 'vl total': vProd, 'vlr': vItemContabil,
                    'icms bc': vBC, 'alq icms': pICMS, 'icms': vICMS, 'ipi': vIPI,
                    'icms st': vICMSST, 'fcp st': vFCPST, 'aql sn': pCredSN, 'icms sn': vCredICMSSN,
                    'descrição': xProd,
                    'COD. PROD.': cProd,
                    'ipi dev': vIPIDevol, 'difal': vICMSUFDest,
                    'COD_PROD_INTERNO': cProd, 'NCM': NCM, 'CFOP': CFOP, 'protocolo': protocolo,
                    'cst_pis': cst_pis, 'vlr_pis': vPIS, 'cst_cofins': cst_cofins, 'vlr_cofins': vCOFINS,
                    'pc': '', 'st': ''
                })

        except Exception as e:
            logging.error(f"Erro ao processar arquivo {arquivo.name}: {e}")
            continue

    return pd.DataFrame(dados)


# -----------------------------
# 2. FUNÇÕES DE EXCEL (ATUALIZADO COM VERMELHO PARA SEM REGRA)
# -----------------------------
def formatar_excel(writer):
    workbook = writer.book

    # --- ESTILOS PADRÃO ---
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    body_font = Font(name='Calibri', size=10)
    border = Border(left=Side(style='thin', color="BFBFBF"), right=Side(style='thin', color="BFBFBF"),
                    top=Side(style='thin', color="BFBFBF"), bottom=Side(style='thin', color="BFBFBF"))

    # --- ESTILOS DE DESTAQUE ---
    highlight_fill_orange = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid') # Laranja (PC)
    highlight_fill_red = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid') # Vermelho (Erro/Sem Regra)
    white_font = Font(name='Calibri', size=10, color='FFFFFF', bold=True) # Fonte branca para fundo vermelho

    for sheet_name in writer.sheets:
        ws = writer.sheets[sheet_name]

        # Formata Cabeçalho
        for cell in ws[1]:
            cell.font = header_font; cell.fill = header_fill; cell.border = border

        # Formata Corpo
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            col_name = str(col[0].value).lower()
            is_money = any(x in col_name for x in ['vlr', 'icms', 'ipi', 'base', 'alq', 'difal', 'total', 'unit', 'operação', 'cont', 'bc'])

            for cell in col[1:]:
                cell.font = body_font; cell.border = border
                if is_money: cell.number_format = '#,##0.00'

                if col_name in ['descrição', 'cfops envolvidos', 'totalizador sete', 'pc']:
                    cell_len = len(str(cell.value)) if cell.value else 0
                    max_len = max(max_len, cell_len)
                elif cell.value:
                    max_len = max(max_len, len(str(cell.value)))

            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 70)
        ws.freeze_panes = 'A2'

        # --- APLICA A COR LARANJA (PC) ---
        if sheet_name in ['Totalizador_PC_CFOP', 'Resumo_Fechado_CFOP']:
            pc_col_idx = None
            for cell in ws[1]:
                if str(cell.value).strip() == 'PC':
                    pc_col_idx = cell.column - 1
                    break

            if pc_col_idx is not None:
                for row in ws.iter_rows(min_row=2):
                    cell_pc = row[pc_col_idx]
                    if str(cell_pc.value) == "PERFUMARIA TC":
                        for cell in row:
                            cell.fill = highlight_fill_orange

        # --- APLICA A COR VERMELHA (SEM REGRA NO RESUMO SETE) ---
        if sheet_name == 'Resumo_SETE_Base':
            totalizador_idx = 0 # Assume primeira coluna 'Totalizador SETE'
            for row in ws.iter_rows(min_row=2):
                cell_totalizador = row[totalizador_idx]
                valor_texto = str(cell_totalizador.value).lower()

                # Se contiver "sem regra específica", pinta de vermelho
                if "sem regra específica" in valor_texto:
                    for cell in row:
                        cell.fill = highlight_fill_red
                        cell.font = white_font

        # --- APLICA A COR VERMELHA (ALERTA PIS COFINS) ---
        if sheet_name == 'Alerta_PIS_COFINS':
             for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.fill = highlight_fill_red
                    cell.font = white_font

def preencher_planilha_sete_existente(df_resumo, caminho_planilha_sete, data_referencia_str):
    if not caminho_planilha_sete or not os.path.exists(caminho_planilha_sete):
        return False, "Caminho inválido."

    try:
        dt = datetime.strptime(data_referencia_str, '%Y-%m-%d')
        nome_aba = dt.strftime('%m.%Y')
        wb = load_workbook(caminho_planilha_sete)

        if nome_aba not in wb.sheetnames:
            return False, f"Aba '{nome_aba}' não encontrada na planilha base."

        ws = wb[nome_aba]

        def get_vals_by_totalizador(nome_totalizador):
            row = df_resumo[df_resumo['Totalizador SETE'] == nome_totalizador]
            if not row.empty:
                return (
                    float(row['VLR OPERAÇÃO'].sum()),
                    float(row['BASE CÁLCULO'].sum()),
                    float(row['VLR ICMS'].sum())
                )
            return (0.0, 0.0, 0.0)

        mapa_linhas = {
            "Saídas Internas Com Benefício":        ("Saídas Internos Com Benefício", 1),
            "Devoluções Internas Com Benefício":    ("Devoluçãos Internos Com Benefício", -1),
            "Saídas Interestaduais Com Benefício":  ("Saídas Interestaduals Com Benefício", 1),
            "Devoluções Interestaduais Com Benefício": ("Devoluçãos Interestaduals Com Benefício", -1)
        }

        linhas_alteradas = 0
        for row in ws.iter_rows(min_row=1, max_row=60, min_col=1, max_col=1):
            cell_val = str(row[0].value).strip().lower() if row[0].value else ""
            for chave_excel, (chave_df, fator) in mapa_linhas.items():
                if chave_excel.lower() in cell_val:
                    valores = get_vals_by_totalizador(chave_df)
                    r = row[0].row
                    ws.cell(row=r, column=2).value = abs(valores[0]) * fator
                    ws.cell(row=r, column=3).value = abs(valores[1]) * fator
                    ws.cell(row=r, column=4).value = abs(valores[2]) * fator
                    linhas_alteradas += 1

        wb.save(caminho_planilha_sete)
        return True, f"Planilha SETE atualizada! ({linhas_alteradas} linhas alteradas)"

    except Exception as e:
        return False, f"Erro ao salvar SETE: {e}"

# -----------------------------
# 3. LÓGICA DE REGRAS E CATEGORIZAÇÃO
# -----------------------------

def carregar_ncms_externos(caminho_arquivo):
    ncms_set = set()
    if not caminho_arquivo or not os.path.exists(caminho_arquivo):
        return ncms_set

    try:
        if str(caminho_arquivo).lower().endswith('.csv'):
            try:
                df = pd.read_csv(caminho_arquivo, header=None, dtype=str, sep=';')
                if df.shape[1] < 5:
                      df = pd.read_csv(caminho_arquivo, header=None, dtype=str, sep=',')
            except:
                df = pd.read_csv(caminho_arquivo, header=None, dtype=str, on_bad_lines='skip')
        else:
            df = pd.read_excel(caminho_arquivo, header=None, dtype=str)

        if df.shape[1] >= 5:
            for index, row in df.iterrows():
                try:
                    col_b_ncm = str(row[1]).strip()
                    col_e_texto = str(row[4]).lower().strip()

                    if 'perfumaria' in col_e_texto:
                        ncm_clean = col_b_ncm.replace('.', '').strip()
                        if ncm_clean:
                            ncms_set.add(ncm_clean)
                except Exception as e_row:
                    continue
        else:
            logging.warning("O arquivo de regras possui menos de 5 colunas (A até E necessárias).")

    except Exception as e:
        logging.error(f"Erro ao ler arquivo de regras: {e}")

    return ncms_set

def definir_invest_simples(row):
    try:
        cod = str(row.get('COD_PROD_INTERNO', '')).upper().strip()
        if cod.startswith('A'):
            return 'SIM'
        return 'NÃO'
    except:
        return 'ERRO'

def definir_nome_totalizador(row):
    invest = row['INVEST']
    cfop = str(row['CFOP']).strip()

    sim_saida_inter = ['6101', '6107', '6401', '6910', '6949', '6911', '6917']
    sim_saida_intra = ['5101', '5401', '5910', '5949']
    sim_dev_inter   = ['2201', '2949', '2410']
    sim_dev_intra   = ['1201', '1410']

    nao_saida_inter = ['6910', '6949', '6102', '6108', '6911', '6917']
    nao_saida_intra = ['5910', '5927', '5102', '5949']
    nao_outros      = ['7101', '7949']
    nao_dev_comp_intra = ['5201', '5556']
    nao_dev_comp_inter = ['6201']
    nao_nao_trib_intra = ['5901', '5915', '5913', '5921']
    nao_nao_trib_inter = ['6901', '6915', '6913', '6921']

    if invest == 'SIM':
        if cfop in sim_saida_inter: return "Saídas Interestaduals Com Benefício"
        if cfop in sim_saida_intra: return "Saídas Internos Com Benefício"
        if cfop in sim_dev_inter:   return "Devoluçãos Interestaduals Com Benefício"
        if cfop in sim_dev_intra:   return "Devoluçãos Internos Com Benefício"

    else:
        if cfop in nao_saida_inter: return "Saídas Interestadual Sem Benefício"
        if cfop in nao_saida_intra: return "Saídas Internos Sem Benefício"
        if cfop in nao_outros:      return "Outros Outros Sem Benefício"
        if cfop in nao_dev_comp_intra: return "OUTROS - DEVOLUÇÃO DE COMPRA INTERNO"
        if cfop in nao_dev_comp_inter: return "OUTROS DEVOLUÇÃO DE COMPRA INTERESTADUAL"
        if cfop in nao_nao_trib_intra: return "OUTROS OUTRAS OPERAÇÃOES NÃO TRIBUTADAS INTERNO"
        if cfop in nao_nao_trib_inter: return "OUTROS OUTRAS OPERAÇÃOES NÃO TRIBUTADAS INTERESTADUAL"

    regiao = 'Interno' if cfop.startswith(('1','5')) else 'Interestadual'
    natureza = 'Devolução' if cfop.startswith(('1','2')) else ('Saída' if cfop.startswith(('5','6')) else 'Outros')
    suffix = "Com Benefício" if invest == 'SIM' else "Sem Benefício"
    # IMPORTANTE: A formatação condicional busca por "Sem Regra Específica"
    return f"{natureza} {regiao} {suffix} (Sem Regra Específica)"

def definir_tipo_pc(row, ncms_validos_set):
    try:
        ncm = str(row.get('NCM', '')).replace('.', '').strip()
        if not ncm: return "(-)"

        if ncm in ncms_validos_set:
            return "PERFUMARIA TC"

        return "(-)"
    except:
        return "(-)"

def verificar_status_pis_cofins(cfop):
    """
    Lista de CFOPs que 'comumente' têm regra de PIS/COFINS (Vendas, Devoluções de Vendas).
    Se o CFOP não estiver aqui, ele vai para o alerta.
    """
    cfops_com_regra = [
        # Saídas Estaduais
        '5101', '5102', '5103', '5104', '5105', '5106', '5109', '5110',
        '5111', '5112', '5113', '5114', '5115', '5116', '5117', '5118',
        '5119', '5120', '5122', '5123', '5124', '5125', '5401', '5402',
        '5403', '5405', '5651', '5652', '5653', '5654', '5655', '5656',

        # Saídas Interestaduais
        '6101', '6102', '6103', '6104', '6105', '6106', '6107', '6108',
        '6109', '6110', '6111', '6112', '6113', '6114', '6115', '6116',
        '6117', '6118', '6119', '6120', '6122', '6123', '6124', '6125',
        '6401', '6402', '6403', '6404', '6651', '6652', '6653', '6654',
        '6655', '6656',

        # Devoluções de Vendas (Entradas)
        '1201', '1202', '1203', '1204', '1410', '1411',
        '2201', '2202', '2203', '2204', '2410', '2411'
    ]

    if str(cfop).strip() in cfops_com_regra:
        return "COM REGRA"
    return "SEM REGRA"

# -----------------------------
# 4. EXECUTOR PRINCIPAL
# -----------------------------
def executar_apuracao_invest(
    pasta_xml: Path,
    caminho_sete: Optional[str] = None,
    caminho_ncm_csv: Optional[str] = None,
    status_callback: Optional[Callable[[str], None]] = None,
    progress_callback: Optional[Callable[[int, int], None]] = None,
    done_callback: Optional[Callable[[str], None]] = None,
    error_callback: Optional[Callable[[str], None]] = None
) -> str:
    logging.info(">>> Iniciando Apuração Invest...")
    if status_callback: status_callback("Iniciando Apuração Invest/Contribuições...")

    ncms_perfumaria_validos = carregar_ncms_externos(caminho_ncm_csv)
    if not ncms_perfumaria_validos:
        logging.warning("Nenhum arquivo de regras carregado ou nenhuma linha com 'perfumaria' encontrada na coluna E.")

    try:
        if status_callback: status_callback("Lendo XMLs...")
        df = ler_xmls_diretamente(pasta_xml, progress_callback)
    except Exception as e:
        logging.error(f"Erro XML: {e}")
        if error_callback: error_callback(f"Erro ao ler XMLs: {e}")
        raise e

    if df.empty:
        msg = "Nenhum dado encontrado nos XMLs."
        if error_callback: error_callback(msg)
        raise ValueError(msg)

    # 3. Processamento
    if status_callback: status_callback("Processando regras e cálculos...")
    df['INVEST'] = df.apply(definir_invest_simples, axis=1)
    df['CFOP_STR'] = df['CFOP'].astype(str).str.strip()
    df['Totalizador SETE'] = df.apply(definir_nome_totalizador, axis=1)
    df['PC'] = df.apply(lambda row: definir_tipo_pc(row, ncms_perfumaria_validos), axis=1)

    # Validação PIS/COFINS
    df['Status_PisCofins'] = df['CFOP_STR'].apply(verificar_status_pis_cofins)

    cols_calc = ['vlr', 'vl total', 'vl unit', 'icms bc', 'icms', 'ipi', 'icms st', 'difal', 'fcp st', 'qnt', 'vlr_pis', 'vlr_cofins']
    for c in cols_calc:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors='coerce').fillna(0.0)

    # Lógica Perfumaria (Zerar valores)
    df_perfumaria = df.copy()
    mask_perfumaria = df_perfumaria['PC'] == 'PERFUMARIA TC'
    cols_to_zero = ['vlr', 'vl total', 'vl unit', 'icms bc', 'icms', 'ipi', 'icms st', 'difal', 'fcp st']
    df_perfumaria.loc[~mask_perfumaria, cols_to_zero] = 0.0
    df_perfumaria['Status Regra'] = df_perfumaria.apply(lambda x: 'NA REGRA' if x['PC'] == 'PERFUMARIA TC' else '(-)', axis=1)

    # 4. Resumos

    # 4.0 Resumo SETE
    resumo_sete = df.groupby(['Totalizador SETE', 'INVEST']).agg(
        vlr=('vlr', 'sum'),
        icms_bc=('icms bc', 'sum'),
        icms=('icms', 'sum'),
        ipi=('ipi', 'sum'),
        difal=('difal', 'sum'),
        icms_st=('icms st', 'sum'),
        cfops=('CFOP_STR', lambda x: ', '.join(sorted(set(x.astype(str)))))
    ).reset_index()

    resumo_sete.rename(columns={
        'vlr': 'VLR OPERAÇÃO',
        'icms_bc': 'BASE CÁLCULO',
        'icms': 'VLR ICMS',
        'ipi': 'VLR IPI',
        'difal': 'VLR DIFAL',
        'icms_st': 'VLR ICMS ST',
        'cfops': 'CFOPs Envolvidos'
    }, inplace=True)

    cols_order = ['Totalizador SETE', 'INVEST', 'VLR OPERAÇÃO', 'BASE CÁLCULO', 'VLR ICMS', 'VLR IPI', 'VLR DIFAL', 'VLR ICMS ST', 'CFOPs Envolvidos']
    resumo_sete = resumo_sete[cols_order]

    # 4.1.A. TOTALIZADOR DETALHADO (PC -> CFOP)
    totalizador_pc = df.pivot_table(
        index=['CFOP', 'PC'],
        values=['vlr', 'icms bc', 'icms', 'ipi', 'icms st', 'difal'],
        aggfunc='sum'
    ).reset_index().sort_values(by=['CFOP', 'PC'])
    totalizador_pc = totalizador_pc[['CFOP', 'PC', 'vlr', 'icms bc', 'icms', 'ipi', 'icms st', 'difal']]
    totalizador_pc.columns = ['CFOP', 'PC', 'VL CONT', 'BC ICMS', 'ICMS', 'IPI', 'ICMS ST', 'DIFAL']

    # 4.1.B. ABA FECHADA
    resumo_fechado_cfop = df.pivot_table(
        index=['CFOP', 'PC'],
        values=['vlr', 'icms bc', 'icms', 'ipi', 'icms st', 'difal'],
        aggfunc='sum'
    ).reset_index().sort_values(by=['CFOP', 'PC'])

    resumo_fechado_cfop = resumo_fechado_cfop[['CFOP', 'PC', 'vlr', 'icms bc', 'icms', 'ipi', 'icms st', 'difal']]
    resumo_fechado_cfop.columns = ['CFOP', 'PC', 'VL CONT', 'BC ICMS', 'ICMS', 'IPI', 'ICMS ST', 'DIFAL']

    # 4.2. Resumo por CST
    resumo_cst = df.pivot_table(index=['cst'], values=['vlr', 'icms bc', 'icms', 'ipi', 'icms st'], aggfunc='sum').reset_index()
    resumo_cst = resumo_cst[['cst', 'vlr', 'icms bc', 'icms', 'ipi', 'icms st']]
    resumo_cst.columns = ['CST/CSOSN', 'Vlr Contábil', 'Base ICMS', 'Vlr ICMS', 'Vlr IPI', 'Vlr ICMS ST']

    # 4.3. ALERTA PIS/COFINS (Nova Aba)
    # Filtra CFOPs que deram "SEM REGRA" e cria um resumo
    df_alerta_pis = df[df['Status_PisCofins'] == 'SEM REGRA'].copy()
    if not df_alerta_pis.empty:
        resumo_alerta_pis = df_alerta_pis.groupby(['CFOP', 'descrição']).agg(
            qnt=('qnt', 'sum'),
            vlr_total=('vlr', 'sum'),
            cst_pis=('cst_pis', 'first'), # Pega o primeiro exemplo
            cst_cofins=('cst_cofins', 'first')
        ).reset_index()
        resumo_alerta_pis.rename(columns={'descrição': 'Exemplo Descrição'}, inplace=True)
    else:
        resumo_alerta_pis = pd.DataFrame(columns=['CFOP', 'Exemplo Descrição', 'qnt', 'vlr_total', 'cst_pis', 'cst_cofins'])

    # 5. Salvar
    if status_callback: status_callback("Gerando Excel...")
    nome_arquivo = f"Resultado_Invest_Contrib_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    caminho_final = pasta_xml / nome_arquivo

    cols_perfumaria_raw = ['n da nf', 'data', 'descrição', 'NCM', 'CFOP', 'Status Regra', 'vlr', 'icms', 'PC']
    cols_perfumaria = [c for c in cols_perfumaria_raw if c in df_perfumaria.columns]

    cols_analise_raw = ['n da nf', 'data', 'descrição', 'NCM', 'CFOP', 'vlr', 'icms', 'PC']
    cols_analise = [c for c in cols_analise_raw if c in df.columns]

    try:
        with pd.ExcelWriter(caminho_final, engine='openpyxl') as writer:
            resumo_sete.to_excel(writer, sheet_name='Resumo_SETE_Base', index=False)

            # Nova Aba de Alerta Vermelho
            resumo_alerta_pis.to_excel(writer, sheet_name='Alerta_PIS_COFINS', index=False)

            df_perfumaria[cols_perfumaria].to_excel(writer, sheet_name='Conferencia_Perfumaria_NCM', index=False)
            df[cols_analise].to_excel(writer, sheet_name='Analise_Geral_PC', index=False)

            totalizador_pc.to_excel(writer, sheet_name='Totalizador_PC_CFOP', index=False)
            resumo_fechado_cfop.to_excel(writer, sheet_name='Resumo_Fechado_CFOP', index=False)
            resumo_cst.to_excel(writer, sheet_name='Resumo_Por_CST', index=False)

            cols_final = ['n da nf', 'cnpj', 'uf', 'data', 'cst', 'qnt', 'vl unit', 'vl total', 'vlr', 'icms bc', 'alq icms', 'icms', 'ipi', 'icms st', 'fcp st', 'aql sn', 'icms sn', 'descrição', 'COD. PROD.', 'ipi dev', 'pc', 'st', 'protocolo', 'difal', 'INVEST', 'Totalizador SETE', 'PC', 'CFOP', 'NCM', 'cst_pis', 'vlr_pis', 'cst_cofins', 'vlr_cofins']
            for c in cols_final:
                if c not in df.columns: df[c] = ''
            df[cols_final].to_excel(writer, sheet_name='Conferencia_Detalhada', index=False)

            formatar_excel(writer)
    except Exception as e:
        logging.error(f"Erro ao salvar Excel: {e}")
        if error_callback: error_callback(f"Erro ao salvar Excel: {e}")
        raise e

    status_msg = ""
    if caminho_sete:
        if status_callback: status_callback("Atualizando Planilha SETE...")
        data_freq = df['data'].mode()[0] if not df.empty else datetime.now().strftime('%Y-%m-%d')
        ok, msg = preencher_planilha_sete_existente(resumo_sete, caminho_sete, data_freq)
        status_msg = f"\n\nStatus Planilha SETE: {msg}"

    if done_callback:
        done_callback(f"{str(caminho_final)}{status_msg}")

    return f"O arquivo foi gerado em: {str(caminho_final)}{status_msg}"
